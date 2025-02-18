import os
import sys
import cv2
import time
import csv
import numpy as np
import mediapipe as mp
import logging
import threading
from datetime import datetime, date
from flask import Flask, Response, render_template_string, request, jsonify, url_for, send_file
from openpyxl import Workbook, load_workbook
import mysql.connector
from html import escape
import re
import pyttsx3  # New TTS engine

# Additional imports for reporting
import io
import base64
import matplotlib.pyplot as plt
from fpdf import FPDF

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

app = Flask(__name__)

# Directories and model paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads", "students_faces")
TTS_FOLDER = os.path.join(BASE_DIR, "static", "tts")
MODEL_YML = os.path.join(BASE_DIR, "train_model.yml")
METADATA_NPZ = os.path.join(BASE_DIR, "model_metadata.npz")
CSV_ATTENDANCE_PATH = os.path.join(BASE_DIR, "attendance.csv")

os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(TTS_FOLDER, exist_ok=True)

# Global variables and lock for thread safety
registration_count = 0         # Count of captured face images (max 5 per registration)
registration_complete = False    # Set to True once training is done
global_lock = threading.Lock()

# If model files exist, set the flag on startup
if os.path.exists(MODEL_YML) and os.path.exists(METADATA_NPZ):
    registration_complete = True
    logging.info("Existing trained model found. Registration complete flag set to True.")

# Global attendance tracking (only one mark per student per day)
# This set will be refreshed at the start of each attendance session by reading today’s CSV records.
attendance_marked_set = set()

# Initialize pyttsx3 engine once
engine = pyttsx3.init()

def speak(text):
    """Speak the provided text in a separate thread."""
    def run():
        try:
            engine.say(text)
            engine.runAndWait()
        except Exception as e:
            logging.error("TTS error: %s", str(e))
    threading.Thread(target=run, daemon=True).start()

# -------------------- Utility Functions --------------------
def sanitize(text):
    """Remove non-alphanumeric characters and extra spaces (replace with underscores)."""
    return re.sub(r'\W+', '_', text.strip())

def load_today_attendance():
    """
    Reload today's attendance records from CSV.
    Returns a set of composite labels (format: "batch_<BatchName>/department_<DepartmentName>/student_<StudentName>")
    """
    today = datetime.now().date()
    records = set()
    if os.path.exists(CSV_ATTENDANCE_PATH):
        with open(CSV_ATTENDANCE_PATH, newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if len(row) < 4:
                    continue
                dept, batch, student, ts = row
                try:
                    record_date = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").date()
                except Exception:
                    continue
                if record_date == today:
                    composite = f"batch_{batch}/department_{dept}/student_{student}"
                    records.add(composite)
    return records

# -------------------- Database Connection --------------------
def get_db_connection():
    try:
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='sems2'
        )
        return connection
    except mysql.connector.Error as err:
        logging.error("Database connection failed: %s", err)
        return None

# -------------------- Endpoints for Dropdown Data --------------------
@app.route('/fetch_batches_departments')
def fetch_batches_departments():
    connection = get_db_connection()
    if connection is None:
        return jsonify({"batches": "", "departments": ""})
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT id, name FROM batches ORDER BY name ASC")
        batches_options = "".join(
            f'<option value="{escape(str(row["id"]))}">{escape(row["name"])}</option>'
            for row in cursor.fetchall()
        )
        cursor.execute("SELECT id, name FROM departments ORDER BY name ASC")
        departments_options = "".join(
            f'<option value="{escape(str(row["id"]))}">{escape(row["name"])}</option>'
            for row in cursor.fetchall()
        )
        cursor.close()
        connection.close()
        return jsonify({"batches": batches_options, "departments": departments_options})
    except Exception as e:
        logging.error("Fetch error: %s", str(e))
        return jsonify({"batches": "", "departments": ""})

@app.route('/fetch_students')
def fetch_students():
    batch_id = request.args.get("batch_id")
    department_id = request.args.get("department_id")
    if not batch_id or not department_id:
        return '<option value="">Invalid parameters</option>'
    connection = get_db_connection()
    if connection is None:
        return '<option value="">Database connection error</option>'
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute(
            "SELECT id, name FROM students WHERE batch_id = %s AND department_id = %s ORDER BY name ASC",
            (batch_id, department_id)
        )
        students = cursor.fetchall()
        cursor.close()
        connection.close()
        if students:
            return "".join(
                f'<option value="{escape(str(s["id"]))}">{escape(s["name"])}</option>' 
                for s in students
            )
        else:
            return '<option value="">No students found</option>'
    except Exception as e:
        logging.error("Error fetching students: %s", str(e))
        return '<option value="">Error fetching students</option>'

# -------------------- ML Training Functions --------------------
def train_face_recognizer():
    """
    Walk through the captured face images stored under UPLOADS_DIR.
    Expected folder structure:
      uploads/students_faces/batch_<BatchName>/department_<DepartmentName>/student_<StudentName>
    The relative folder path is used as the composite label.
    """
    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
    except Exception as e:
        logging.error("Error creating LBPH face recognizer: %s", str(e))
        return None, None

    faces = []
    labels = []
    label_map = {}  # Mapping: composite label -> integer
    current_label = 0

    for root, dirs, files in os.walk(UPLOADS_DIR):
        for file in files:
            if file.lower().endswith(".jpg"):
                path = os.path.join(root, file)
                img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
                if img is None:
                    continue
                composite_label = os.path.relpath(root, UPLOADS_DIR)
                if composite_label not in label_map:
                    label_map[composite_label] = current_label
                    current_label += 1
                label = label_map[composite_label]
                faces.append(img)
                labels.append(label)

    if len(faces) == 0:
        logging.error("No face images found for training.")
        return None, None

    recognizer.train(faces, np.array(labels))
    recognizer.write(MODEL_YML)
    np.savez(METADATA_NPZ, label_map=label_map)
    logging.info("Training complete. Model saved with %d labels.", len(label_map))
    return recognizer, label_map

def train_model_process():
    global registration_complete
    logging.info("Training started...")
    recognizer, label_map = train_face_recognizer()
    if recognizer is not None:
        with global_lock:
            registration_complete = True
        logging.info("Model training complete.")
    else:
        logging.error("Model training failed due to lack of data.")

# -------------------- CSV Attendance Function --------------------
def mark_attendance(composite_label):
    """
    Mark attendance by appending a record to the CSV file.
    The composite label is in the format "batch_<BatchName>/department_<DepartmentName>/student_<StudentName>".
    Extract the actual names for storage.
    """
    parts = os.path.normpath(composite_label).split(os.sep)
    if len(parts) != 3:
        logging.error("Unexpected label format: %s", composite_label)
        return
    batch_part, dept_part, student_part = parts
    batch_name = batch_part.replace("batch_", "")
    dept_name = dept_part.replace("department_", "")
    student_name = student_part.replace("student_", "")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(CSV_ATTENDANCE_PATH, mode="a", newline="") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([dept_name, batch_name, student_name, timestamp])
    attendance_marked_set.add(composite_label)
    logging.info("Attendance marked for %s at %s", composite_label, timestamp)

# -------------------- Download CSV Endpoint --------------------
@app.route("/download_attendance")
def download_attendance():
    if os.path.exists(CSV_ATTENDANCE_PATH):
        return send_file(CSV_ATTENDANCE_PATH, as_attachment=True)
    else:
        return "No attendance records available", 404

# -------------------- UI Templates (Apple Inspired Minimalism) --------------------
navbar = """
<nav class="navbar navbar-expand-lg navbar-light bg-white border-bottom" style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;">
  <div class="container">
    <a class="navbar-brand text-dark" href="{{ url_for('index') }}"><i class="fas fa-school"></i> FaceAttend</a>
    <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
      <span class="navbar-toggler-icon"></span>
    </button>
    <div class="collapse navbar-collapse" id="navbarNav">
      <ul class="navbar-nav ms-auto">
        <li class="nav-item"><a class="nav-link text-dark" href="{{ url_for('register') }}">Register</a></li>
        <li class="nav-item"><a class="nav-link text-dark" href="{{ url_for('attendance') }}">Attendance</a></li>
        <li class="nav-item"><a class="nav-link text-dark" href="{{ url_for('report') }}">Reports</a></li>
      </ul>
    </div>
  </div>
</nav>
"""

footer = """
<footer class="text-center mt-5 py-3" style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;">
  <small>&copy; 2025 FaceAttend. All Rights Reserved.</small>
</footer>
"""

home_template = """
<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>FaceAttend - Home</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/dist/css/all.min.css" rel="stylesheet">
    <style>
      body { background: #ffffff; }
      .hero { background: url('https://source.unsplash.com/1600x400/?technology,apple') center/cover; color: #333; padding: 5rem 2rem; }
      .hero h1 { font-size: 3rem; font-weight: 600; }
      .card { border: none; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
    </style>
  </head>
  <body>
    """ + navbar + """
    <div class="hero text-center">
      <h1>Welcome to FaceAttend</h1>
      <p>Your smart face recognition attendance system</p>
      <a class="btn btn-outline-dark btn-lg me-2" href="{{ url_for('register') }}">Register Face</a>
      <a class="btn btn-dark btn-lg" href="{{ url_for('attendance') }}">Mark Attendance</a>
    </div>
    <div class="container my-5">
      <div class="row text-center">
         <div class="col-md-4">
           <div class="card p-4 mb-4">
             <i class="fas fa-user-plus fa-3x mb-3 text-primary"></i>
             <h5 class="card-title">Easy Registration</h5>
             <p class="card-text">Register your face quickly with a seamless experience.</p>
           </div>
         </div>
         <div class="col-md-4">
           <div class="card p-4 mb-4">
             <i class="fas fa-check fa-3x mb-3 text-success"></i>
             <h5 class="card-title">Real-Time Attendance</h5>
             <p class="card-text">Instantly mark your attendance with cutting-edge recognition.</p>
           </div>
         </div>
         <div class="col-md-4">
           <div class="card p-4 mb-4">
             <i class="fas fa-chart-line fa-3x mb-3 text-info"></i>
             <h5 class="card-title">Analytics & Reports</h5>
             <p class="card-text">Review detailed reports with intuitive visuals.</p>
           </div>
         </div>
      </div>
      <div class="text-center">
        <a class="btn btn-outline-secondary" href="{{ url_for('download_attendance') }}"><i class="fas fa-download"></i> Download Attendance CSV</a>
      </div>
    </div>
    """ + footer + """
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
  </body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(home_template)

register_template = """
<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>FaceAttend - Registration</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/dist/css/all.min.css" rel="stylesheet">
    <style>
      body { background: #ffffff; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; }
      .card { margin-top: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
      img { width: 100%; max-width: 600px; border: 1px solid #ddd; }
    </style>
  </head>
  <body>
    """ + navbar + """
    <div class="container">
      <div class="card">
        <div class="card-header bg-primary text-white">
          <h4><i class="fas fa-user-plus"></i> Face Registration Module</h4>
        </div>
        <div class="card-body">
          <!-- Reset registration count for new sessions -->
          <script>window.onload = () => { fetch("{{ url_for('reset_registration') }}"); };</script>
          <form id="registrationForm">
            <div class="row mb-3">
              <div class="col-md-4">
                <label for="batchSelect" class="form-label">Select Batch</label>
                <select class="form-select" id="batchSelect"></select>
              </div>
              <div class="col-md-4">
                <label for="departmentSelect" class="form-label">Select Department</label>
                <select class="form-select" id="departmentSelect"></select>
              </div>
              <div class="col-md-4">
                <label for="studentSelect" class="form-label">Select Student</label>
                <select class="form-select" id="studentSelect"></select>
              </div>
            </div>
            <button type="button" id="captureBtn" class="btn btn-primary"><i class="fas fa-camera"></i> Capture Faces</button>
            <button type="button" id="trainBtn" class="btn btn-success" disabled><i class="fas fa-cogs"></i> Train Model</button>
            <p id="statusMsg" class="mt-3"></p>
          </form>
          <div class="mt-4 text-center">
            <img id="videoFeed" src="" alt="Video Feed">
          </div>
        </div>
      </div>
    </div>
    """ + footer + """
    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script>
      $(document).ready(function(){
          $.getJSON("{{ url_for('fetch_batches_departments') }}", function(data){
              $("#batchSelect").html(data.batches);
              $("#departmentSelect").html(data.departments);
          });
          $("#batchSelect, #departmentSelect").change(function(){
              var batch_id = $("#batchSelect").val();
              var department_id = $("#departmentSelect").val();
              $.get("{{ url_for('fetch_students') }}", {batch_id: batch_id, department_id: department_id}, function(data){
                  $("#studentSelect").html(data);
              });
          });
          
          $("#captureBtn").click(function(){
              var batch_id = $("#batchSelect").val();
              var department_id = $("#departmentSelect").val();
              var student_id = $("#studentSelect").val();
              var student_name = $("#studentSelect option:selected").text();
              var batch_name = $("#batchSelect option:selected").text();
              var department_name = $("#departmentSelect option:selected").text();
              if(!batch_id || !department_id || !student_id){
                  $("#statusMsg").text("Please select batch, department, and student.");
                  return;
              }
              $("#statusMsg").text("Capturing faces... Please wait until 5 images are captured.");
              var feedUrl = "{{ url_for('video_feed_register') }}?batch_id=" + encodeURIComponent(batch_id) +
                            "&department_id=" + encodeURIComponent(department_id) +
                            "&student_id=" + encodeURIComponent(student_id) +
                            "&student_name=" + encodeURIComponent(student_name) +
                            "&batch_name=" + encodeURIComponent(batch_name) +
                            "&department_name=" + encodeURIComponent(department_name);
              $("#videoFeed").attr("src", feedUrl);
          });
          
          setInterval(function(){
              $.getJSON("{{ url_for('registration_status') }}", function(data){
                  if(data.registration_count >= 5){
                      $("#statusMsg").text("Face capture complete. Click 'Train Model' to proceed.");
                      $("#trainBtn").prop("disabled", false);
                  }
              });
          }, 1000);
          
          $("#trainBtn").click(function(){
              $(this).prop("disabled", true);
              $("#statusMsg").text("Training model... Please wait.");
              $.post("{{ url_for('train_model') }}", function(response){
                  $("#statusMsg").text(response.message);
              }, "json");
          });
      });
    </script>
  </body>
</html>
"""

@app.route("/register")
def register():
    return render_template_string(register_template)

attendance_template = """
<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>FaceAttend - Attendance</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0-beta3/css/all.min.css" rel="stylesheet">
    <style>
      body { background: #ffffff; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; }
      .card { margin-top: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
      img { width: 100%; max-width: 600px; border: 1px solid #ddd; }
    </style>
  </head>
  <body>
    """ + navbar + """
    <div class="container">
      <div class="card">
        <div class="card-header bg-success text-white">
          <h4><i class="fas fa-check"></i> Attendance Module</h4>
        </div>
        <div class="card-body text-center">
          <p>Point your face towards the camera to mark attendance.</p>
          <div class="mt-4">
            <img id="videoFeedAtt" src="" alt="Attendance Feed">
          </div>
          <p id="attendanceMsg" class="mt-3"></p>
        </div>
      </div>
    </div>
    """ + footer + """
    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script>
       $(document).ready(function(){
           $("#videoFeedAtt").attr("src", "{{ url_for('video_feed_attendance') }}");
       });
    </script>
  </body>
</html>
"""

@app.route("/attendance")
def attendance():
    return render_template_string(attendance_template)

# -------------------- Report Generation --------------------
def generate_report_data(batch_id, department_id):
    """
    Generates report data by:
      - Querying the DB for batch, department, and students.
      - Reading the CSV attendance file (records are stored as: dept, batch, student, timestamp).
      - For the given batch and department, computing for each student the unique attendance dates.
    Returns a dictionary with batch name, department name, a list of students, report_data (list of dicts),
    total working days, and a mapping of student name to set of attendance dates.
    """
    connection = get_db_connection()
    if connection is None:
         return None
    cursor = connection.cursor(dictionary=True)
    cursor.execute("SELECT name FROM batches WHERE id = %s", (batch_id,))
    batch_row = cursor.fetchone()
    batch_name = batch_row["name"] if batch_row else "Unknown"
    cursor.execute("SELECT name FROM departments WHERE id = %s", (department_id,))
    dept_row = cursor.fetchone()
    dept_name = dept_row["name"] if dept_row else "Unknown"
    cursor.execute("SELECT id, name FROM students WHERE batch_id = %s AND department_id = %s ORDER BY name ASC", (batch_id, department_id))
    students = cursor.fetchall()
    cursor.close()
    connection.close()
    attendance_data = []
    if os.path.exists(CSV_ATTENDANCE_PATH):
         with open(CSV_ATTENDANCE_PATH, newline='') as csvfile:
              reader = csv.reader(csvfile)
              for row in reader:
                   if len(row) < 4:
                        continue
                   row_dept, row_batch, row_student, ts = row
                   if row_batch == batch_name and row_dept == dept_name:
                        try:
                             date_obj = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").date()
                        except Exception:
                             continue
                        attendance_data.append((row_student, date_obj))
    working_days = set()
    student_attendance = {s["name"]: set() for s in students}
    for row_student, date_obj in attendance_data:
         working_days.add(date_obj)
         if row_student in student_attendance:
              student_attendance[row_student].add(date_obj)
    total_working_days = len(working_days) if len(working_days) > 0 else 1
    report_data = []
    for s in students:
         name = s["name"]
         att_count = len(student_attendance.get(name, set()))
         perc = (att_count / total_working_days) * 100
         absent_count = total_working_days - att_count
         absent_perc = (absent_count / total_working_days) * 100
         report_data.append({
             "name": name,
             "attendance_count": att_count,
             "attendance_percentage": round(perc, 2),
             "absent_count": absent_count,
             "absent_percentage": round(absent_perc, 2)
         })
    return {
         "batch_name": batch_name,
         "dept_name": dept_name,
         "students": students,
         "report_data": report_data,
         "total_working_days": total_working_days,
         "student_attendance": student_attendance
    }

# A simple filter form template for report generation
report_form_template = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Generate Attendance Report</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
</head>
<body>
<div class="container mt-5">
  <h2>Generate Attendance Report</h2>
  <form method="get" action="/report">
    <div class="mb-3">
      <label for="batchSelect" class="form-label">Select Batch</label>
      <select class="form-select" id="batchSelect" name="batch_id"></select>
    </div>
    <div class="mb-3">
      <label for="departmentSelect" class="form-label">Select Department</label>
      <select class="form-select" id="departmentSelect" name="department_id"></select>
    </div>
    <button type="submit" class="btn btn-primary">Generate Report</button>
  </form>
</div>
<script>
  $(document).ready(function(){
    $.getJSON("/fetch_batches_departments", function(data){
      $("#batchSelect").html(data.batches);
      $("#departmentSelect").html(data.departments);
    });
  });
</script>
</body>
</html>
"""

# Template to display the report with embedded graphs and download links
report_template = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Attendance Report</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    .graph { max-width: 600px; margin: 20px auto; }
  </style>
</head>
<body>
<div class="container mt-5">
  <h2>Attendance Report for Batch: {{ batch_name }} | Department: {{ dept_name }}</h2>
  <p>Total Working Days: {{ total_working_days }}</p>
  <table class="table table-bordered">
    <thead>
      <tr>
        <th>Student Name</th>
        <th>Attendance Count</th>
        <th>Attendance Percentage</th>
        <th>Absent Count</th>
        <th>Absent Percentage</th>
      </tr>
    </thead>
    <tbody>
      {% for row in report_data %}
      <tr>
        <td>{{ row.name }}</td>
        <td>{{ row.attendance_count }}</td>
        <td>{{ row.attendance_percentage }}%</td>
        <td>{{ row.absent_count }}</td>
        <td>{{ row.absent_percentage }}%</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  <div class="graph text-center">
    <h4>Attendance Percentage per Student</h4>
    <img src="data:image/png;base64,{{ bar_graph }}" class="img-fluid">
  </div>
  <div class="graph text-center">
    <h4>Overall Attendance</h4>
    <img src="data:image/png;base64,{{ pie_chart }}" class="img-fluid">
  </div>
  <div class="mt-4 text-center">
    <a href="{{ url_for('download_excel_report', batch_id=batch_id, department_id=department_id) }}" class="btn btn-success">Download Excel Report</a>
    <a href="{{ url_for('download_pdf_report', batch_id=batch_id, department_id=department_id) }}" class="btn btn-danger">Download PDF Report</a>
  </div>
</div>
</body>
</html>
"""

@app.route("/report")
def report():
    batch_id = request.args.get("batch_id")
    department_id = request.args.get("department_id")
    if not batch_id or not department_id:
         return render_template_string(report_form_template)
    data = generate_report_data(batch_id, department_id)
    if data is None:
         return "Error generating report data"
    # Generate bar graph: Attendance percentage per student
    plt.figure(figsize=(10,6))
    names = [d["name"] for d in data["report_data"]]
    percentages = [d["attendance_percentage"] for d in data["report_data"]]
    plt.bar(names, percentages, color='green')
    plt.xlabel("Students")
    plt.ylabel("Attendance Percentage")
    plt.title("Attendance Percentage per Student")
    plt.xticks(rotation=45, ha="right")
    buf_bar = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buf_bar, format='png')
    buf_bar.seek(0)
    bar_graph = base64.b64encode(buf_bar.getvalue()).decode('utf-8')
    plt.close()

    # Generate pie chart: Overall attendance vs absent
    total_possible = len(data["students"]) * data["total_working_days"]
    total_attended = sum(len(v) for v in data["student_attendance"].values())
    overall_attendance_percentage = (total_attended / total_possible) * 100 if total_possible > 0 else 0
    overall_absent_percentage = 100 - overall_attendance_percentage
    plt.figure(figsize=(6,6))
    plt.pie([overall_attendance_percentage, overall_absent_percentage], labels=["Present", "Absent"], autopct='%1.1f%%', colors=["green", "red"])
    plt.title("Overall Attendance")
    buf_pie = io.BytesIO()
    plt.savefig(buf_pie, format='png')
    buf_pie.seek(0)
    pie_chart = base64.b64encode(buf_pie.getvalue()).decode('utf-8')
    plt.close()

    return render_template_string(report_template,
                                  batch_id=batch_id,
                                  department_id=department_id,
                                  batch_name=data["batch_name"],
                                  dept_name=data["dept_name"],
                                  report_data=data["report_data"],
                                  total_working_days=data["total_working_days"],
                                  bar_graph=bar_graph,
                                  pie_chart=pie_chart)

@app.route("/download_excel_report")
def download_excel_report():
    batch_id = request.args.get("batch_id")
    department_id = request.args.get("department_id")
    data = generate_report_data(batch_id, department_id)
    if data is None:
         return "Error generating report data"
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    headers = ["Student Name", "Attendance Count", "Attendance Percentage", "Absent Count", "Absent Percentage"]
    ws.append(headers)
    for row in data["report_data"]:
         ws.append([row["name"], row["attendance_count"], row["attendance_percentage"], row["absent_count"], row["absent_percentage"]])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="attendance_report.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/download_pdf_report")
def download_pdf_report():
    batch_id = request.args.get("batch_id")
    department_id = request.args.get("department_id")
    data = generate_report_data(batch_id, department_id)
    if data is None:
         return "Error generating report data"
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Attendance Report", ln=1, align="C")
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Batch: {data['batch_name']} | Department: {data['dept_name']}", ln=1, align="C")
    pdf.ln(5)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(40, 10, "Student Name", 1)
    pdf.cell(40, 10, "Attendance Count", 1)
    pdf.cell(40, 10, "Attendance %", 1)
    pdf.cell(30, 10, "Absent Count", 1)
    pdf.cell(30, 10, "Absent %", 1)
    pdf.ln()
    pdf.set_font("Arial", "", 10)
    for row in data["report_data"]:
         pdf.cell(40, 10, row["name"], 1)
         pdf.cell(40, 10, str(row["attendance_count"]), 1)
         pdf.cell(40, 10, f"{row['attendance_percentage']}%", 1)
         pdf.cell(30, 10, str(row["absent_count"]), 1)
         pdf.cell(30, 10, f"{row['absent_percentage']}%", 1)
         pdf.ln()
    pdf_str = pdf.output(dest='S').encode('latin1')
    output = io.BytesIO(pdf_str)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="attendance_report.pdf", mimetype="application/pdf")

# -------------------- Video Streaming Endpoints --------------------
@app.route("/video_feed_register")
def video_feed_register():
    global registration_count
    batch_id = request.args.get("batch_id", "unknown")
    department_id = request.args.get("department_id", "unknown")
    student_id = request.args.get("student_id", "unknown")
    student_name = request.args.get("student_name", "unknown")
    batch_name = request.args.get("batch_name", "unknown")
    department_name = request.args.get("department_name", "unknown")
    safe_batch = "batch_" + sanitize(batch_name)
    safe_dept = "department_" + sanitize(department_name)
    safe_student = "student_" + sanitize(student_name)
    dest_dir = os.path.join(UPLOADS_DIR, safe_batch, safe_dept, safe_student)
    os.makedirs(dest_dir, exist_ok=True)
    
    # Reset the registration count for new sessions
    with global_lock:
        registration_count = 0

    def gen():
        global registration_count
        cap = cv2.VideoCapture(0)
        mp_face = mp.solutions.face_detection.FaceDetection(min_detection_confidence=0.5)
        try:
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret or frame is None or frame.shape[0] == 0 or frame.shape[1] == 0:
                    logging.error("Empty frame in registration; skipping...")
                    continue
                try:
                    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                except Exception as e:
                    logging.error("Error converting frame color: %s", str(e))
                    continue
                results = mp_face.process(rgb_frame)
                with global_lock:
                    current_count = registration_count
                if results.detections:
                    detection = results.detections[0]
                    bboxC = detection.location_data.relative_bounding_box
                    ih, iw, _ = frame.shape
                    x = int(bboxC.xmin * iw)
                    y = int(bboxC.ymin * ih)
                    w = int(bboxC.width * iw)
                    h = int(bboxC.height * ih)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
                    with global_lock:
                        if registration_count < 5:
                            face_img = frame[y:y+h, x:x+w]
                            timestamp = int(time.time())
                            filename = os.path.join(dest_dir, f"face_{registration_count}_{timestamp}.jpg")
                            cv2.imwrite(filename, face_img)
                            registration_count += 1
                            cv2.putText(frame, f"Captured {registration_count}/5", (x, y-10),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)
                            time.sleep(0.5)
                else:
                    cv2.putText(frame, "No face detected", (20, 30),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                with global_lock:
                    if registration_count >= 5:
                        cv2.putText(frame, "Capture complete. Please train the model.", (20, 60),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
                ret, jpeg = cv2.imencode('.jpg', frame)
                frame_bytes = jpeg.tobytes()
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
                with global_lock:
                    if registration_count >= 5:
                        break
        except Exception as e:
            logging.error("Error in video_feed_register: %s", str(e))
        finally:
            cap.release()
    return Response(gen(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/registration_status')
def registration_status():
    with global_lock:
        count = registration_count
        complete = registration_complete
    return jsonify({"registration_count": count, "registration_complete": complete})

@app.route('/train_model', methods=['POST'])
def train_model():
    global registration_count, registration_complete
    with global_lock:
        if registration_count < 5:
            return jsonify({"message": "Not enough face captures to train the model."})
    threading.Thread(target=train_model_process).start()
    return jsonify({"message": "Training initiated. Please wait..."})

@app.route("/video_feed_attendance")
def video_feed_attendance():
    global registration_complete, attendance_marked_set
    # Refresh today’s attendance records from CSV so that a student is only marked once per day even after a restart
    attendance_marked_set = load_today_attendance()

    # Use the model if already trained or if model files exist
    if registration_complete or (os.path.exists(MODEL_YML) and os.path.exists(METADATA_NPZ)):
        try:
            recognizer = cv2.face.LBPHFaceRecognizer_create()
            recognizer.read(MODEL_YML)
            metadata = np.load(METADATA_NPZ, allow_pickle=True)
            label_map = metadata['label_map'].item()
            inv_label_map = {v: k for k, v in label_map.items()}
        except Exception as e:
            logging.error("Error loading trained model: %s", str(e))
            recognizer = None
    else:
        recognizer = None

    RECOGNITION_THRESHOLD = 80

    def gen():
        cap = cv2.VideoCapture(0)
        mp_face = mp.solutions.face_detection.FaceDetection(min_detection_confidence=0.5)
        try:
            while cap.isOpened():
                ret, frame = cap.read()
                if not ret or frame is None or frame.shape[0]==0 or frame.shape[1]==0:
                    logging.error("Empty frame in attendance; skipping...")
                    continue
                try:
                    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                except Exception as e:
                    logging.error("Error converting frame in attendance: %s", str(e))
                    continue
                results = mp_face.process(rgb_frame)
                if results.detections:
                    for detection in results.detections:
                        bboxC = detection.location_data.relative_bounding_box
                        ih, iw, _ = frame.shape
                        x = int(bboxC.xmin * iw)
                        y = int(bboxC.ymin * ih)
                        w = int(bboxC.width * iw)
                        h = int(bboxC.height * ih)
                        if recognizer is not None:
                            face_roi = frame[y:y+h, x:x+w]
                            try:
                                gray_face = cv2.cvtColor(face_roi, cv2.COLOR_BGR2GRAY)
                            except Exception as e:
                                logging.error("Error converting face ROI: %s", str(e))
                                continue
                            try:
                                label_pred, confidence = recognizer.predict(gray_face)
                            except Exception as e:
                                logging.error("Prediction error: %s", str(e))
                                continue
                            predicted_composite = inv_label_map.get(label_pred, "Unknown")
                            if confidence < RECOGNITION_THRESHOLD:
                                parts = predicted_composite.split(os.sep)
                                if len(parts) == 3:
                                    batch_name = parts[0].replace("batch_", "")
                                    dept_name = parts[1].replace("department_", "")
                                    student_name = parts[2].replace("student_", "")
                                else:
                                    batch_name = dept_name = student_name = "Unknown"
                                if predicted_composite not in attendance_marked_set:
                                    mark_attendance(predicted_composite)
                                    status_text = f"Attendance Marked: {student_name}"
                                    speak(status_text)
                                else:
                                    status_text = f"Already Marked: {student_name}"
                                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                                cv2.putText(frame, status_text, (x, y-30),
                                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                                cv2.putText(frame, f"{batch_name}, {dept_name}", (x, y-10),
                                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                            else:
                                cv2.putText(frame, "Face Not Recognized", (x, y-30),
                                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
                        else:
                            cv2.putText(frame, "Model Not Trained", (x, y-30),
                                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 0, 0), 2)
                            cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
                ret, jpeg = cv2.imencode('.jpg', frame)
                frame_bytes = jpeg.tobytes()
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
        except Exception as e:
            logging.error("Error in video_feed_attendance: %s", str(e))
        finally:
            cap.release()
    return Response(gen(), mimetype='multipart/x-mixed-replace; boundary=frame')

# -------------------- Endpoint to Reset Registration Count --------------------
@app.route("/reset_registration")
def reset_registration():
    global registration_count
    with global_lock:
        registration_count = 0
    return "Registration counter reset", 200

if __name__ == "__main__":
    app.run(debug=True)
